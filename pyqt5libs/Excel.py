# coding=utf-8
import contextlib
import os
from datetime import date

import xlsxwriter
from openpyxl import load_workbook
from xlsxwriter.utility import xl_rowcol_to_cell

try:
    import win32com.client as win32

    WIN32 = True
except ImportError:
    WIN32 = False

from . import Ventanas
from .utiles import saveFileDialog, AbrirArchivo, FormatoFecha


class Excel:
    archivo = ''  # nombre de archivo
    libro = None
    hoja = None
    cabeceras = {}
    
    hojas = {}
    
    formato = None
    formatos = {
        'moneda': {'num_format': '$#,##0.00'},
        'negrita': {'bold': True},
        'centradoh': {'align': 'center'},
        'porcentaje': {'num_format': '0.00%'},
        'numero': {'num_format': '#,##0.00'},
        'fecha': {'num_format': 'dd/mm/yyyy'},
        'fechahora': {'num_format': 'dd/mm/yyyy hh:mm:ss'},
    }
    max_row = 0  # ultima fila de un archivo

    def __init__(self, archivo='', hoja=None):
        self.hojas = {}
        self.archivo = archivo
        self.libro: xlsxwriter.Workbook = None
    
    def asegura_libro(self):
        """Inicializa el libro si aún no existe."""
        if not self.libro:
            self.libro = xlsxwriter.Workbook(self.archivo)
            print(f"Libro creado: {self.archivo}")
    
    def ArmaCabeceras(self, cabeceras, fila=0, col_inicio=0, formato=None):
        if formato:
            formato_celda = self.libro.add_format(formato)
        else:
            formato_celda = self.libro.add_format({
                'bold': True,
                'border': 2,
                'align': 'center',
                'bg_color': '#D9D9D9'  # Gris claro
            })
        if not self.archivo:
            return

        if isinstance(cabeceras, list):
            # Diccionario con { "nombre_cabecera": número_de_columna }
            cabeceras = {cabecera: i for i, cabecera in enumerate(cabeceras)}

        for k, v in cabeceras.items():
            self.hoja.write(fila, v + col_inicio, k, formato_celda)

        self.cabeceras = cabeceras

    def ObtieneArchivo(self, archivo: str = 'excel/archivo.xlsx') -> str:
        self.archivo = saveFileDialog(filename=archivo, files="Archivos de Excel (*.xlsx)")

        # if self.archivo and not self.archivo.upper().endswith("XLSX"):
        #     self.archivo += ".XLSX"

        if self.archivo:
            self.libro = xlsxwriter.Workbook(self.archivo)
            self.hoja = self.crea_hoja()

        return self.archivo

    def crea_hoja(self, nombre_hoja: str = 'Hoja1'):
        self.agrega_hoja(nombre_hoja)
    
    def agrega_hoja(self, nombre_hoja: str = 'Hoja1'):
        self.asegura_libro()
        nombre_hoja = nombre_hoja[:30].translate(str.maketrans("áéíóúÁÉÍÓÚñÑ", "aeiouAEIOUnN"))
        if nombre_hoja in self.hojas:
            print(f"La hoja '{nombre_hoja}' ya existe.")
            self.hoja = self.hojas[nombre_hoja]
            return True
        else:
            self.hoja = self.libro.add_worksheet(nombre_hoja)
            self.hojas[nombre_hoja] = self.hoja # Guarda la referencia
            return False

    def activa_hoja(self, nombre_hoja: str = 'Hoja1'):
        """Activa una hoja existente o crea una nueva si no existe."""
        if nombre_hoja in self.hojas:
            self.hoja = self.hojas[nombre_hoja]
            return True
        else:
            return self.agrega_hoja(nombre_hoja)
    
    def Titulo(self, titulo: str = '', desdecol: str = 'A', hastacol: str = 'A',
               fila: int = 0, combina: bool = True, **kwargs) -> None:
        # Create a format to use in the merged range.
        merge_format = self.libro.add_format({
            'bold': 1,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
        })
        if 'tamanio' in kwargs:
            merge_format.set_size(kwargs['tamanio'])
        else:
            merge_format.set_size(14)
        if combina:
            self.hoja.merge_range('{}{}:{}{}'.format(desdecol, fila + 1, hastacol, fila + 1), titulo, merge_format)
        else:
            self.hoja.write('{}{}'.format(fila, desdecol))

    def AgregarTabla(self, table='', header=[]):
        if table:
            self.hoja.add_table(table, {'columns': header})

    def Cerrar(self, abre=True, autofilter='', table='', header=[]):
        if autofilter:
            self.hoja.autofilter(autofilter)
        if table:
            self.hoja.add_table(table, {'columns': header})
        try:
            if not WIN32:
                for v in self.cabeceras.values():
                    self.set_column_autowidth(v)
            self.libro.close()
            if WIN32:
                xlautofit(self.archivo)

        except IOError:
            Ventanas.showAlert("Sistema", "No se puede escribir el archivo {} esta abierto. Intente cerrarlo".format(
                self.archivo
            ))

        if abre:
            AbrirArchivo(self.archivo)

    def EscribeFila(self, datos='', fila=1, inicio=0, formato=None, color_fondo=None, **kwargs):
        col = inicio
        for d in datos:
            formato_celda = None

            if isinstance(d, bool):
                item = 'SI' if d else 'NO'
            else:
                item = d

            # Construcción del formato
            if formato and col in formato:
                formato_base = formato[col]
            else:
                formato_base = {}

            if color_fondo:
                formato_base = dict(formato_base)  # Copia para no alterar el original
                formato_base.update({'bg_color': color_fondo})

            if formato_base:
                formato_celda = self.libro.add_format(formato_base)

            if formato_celda:
                self.hoja.write(fila, col, item, formato_celda)
            else:
                self.hoja.write(fila, col, item)

            col += 1

    def Totales(self, columnas, desdefila, hastafila, filaformula, formato=None):
        for col in columnas:
            formato_celda = None
            if formato:
                if isinstance(formato, dict):
                    formato_celda = self.libro.add_format(formato)
                else:
                    formato_celda = formato
            if formato_celda:
                self.hoja.write_formula(filaformula, col, '=sum({}{}:{}{})'.format(
                    chr(col + 65), desdefila, chr(col + 65), hastafila
                ), formato_celda)
            else:
                self.hoja.write_formula(filaformula, col, '=sum({}{}:{}{})'.format(
                    chr(col + 65), desdefila, chr(col + 65), hastafila
                ))

    def SubTitulo(self, titulo='', desdecol='A', hastacol='A', fila=0, combina=True):
        self.Titulo(titulo, desdecol, hastacol, fila, combina, tamanio=12)

    def get_column_width(self, column: int):
        """Get the max column width in a `Worksheet` column."""
        worksheet = self.hoja
        strings = getattr(worksheet, '_ts_all_strings', None)
        if strings is None:
            strings = worksheet._ts_all_strings = sorted(
                worksheet.str_table.string_table,
                key=worksheet.str_table.string_table.__getitem__)
        lengths = set()
        for row in range(worksheet.dim_rowmax + 1):
            try:
                cell = worksheet.table.get(row, {}).get(column)
                if cell is not None:
                    value = str(cell)
                    iter_length = len(value)
                    if iter_length:
                        lengths.add(iter_length)
            except Exception:
                continue
        if not lengths:
            return None
        return max(lengths)

    def set_column_autowidth(self, column: int):
        """
        Set the width automatically on a column in the `Worksheet`.
        !!! Make sure you run this function AFTER having all cells filled in
        the worksheet!
        """
        maxwidth = self.get_column_width(column=column)
        if maxwidth is None:
            return
        self.hoja.set_column(first_col=column, last_col=column, width=maxwidth)

    def EscribeFilaColumna(self, fila: int = 0, columna=0, valor=None, formato=None, combina=None):
        if formato:
            if isinstance(formato, dict):
                formato_celda = self.libro.add_format(formato)
            else:
                formato_celda = formato
            if combina:
                self.hoja.merge_range(fila, columna, combina[0], combina[1], valor, formato_celda)
            else:
                if isinstance(columna, str):
                    self.hoja.write(f'{columna}{fila}', valor, formato_celda)
                else:
                    self.hoja.write(fila, columna, valor, formato_celda)
        else:
            if combina:
                self.hoja.merge_range(fila, columna, combina[0], combina[1], valor)
            else:
                if isinstance(columna, str):
                    self.hoja.write(f'{columna}{fila}', valor)
                else:
                    self.hoja.write(fila, columna, valor)

    def AgregaFormato(self):
        self.formato = None
        self.formato = self.libro.add_format()
        return self.formato

    def FormatoNumero(self, formato_numero):
        self.formato.set_num_format(formato_numero)

    def FormatoNegrita(self, negrita=True):
        self.formato.set_bold(negrita)

    def EstableceEncabezado(self, encabezados: dict, imagen='', alineacionimagen='L'):
        alineacionesimagenes = {
            'L': 'image_left',
            'C': 'image_center',
            'R': 'image_right'
        }
        texto = ''
        for k, v in encabezados.items():
            texto += f'&{v}{k}'
        if imagen:
            texto += f'&{alineacionimagen}&[Picture]'
            opciones = {alineacionesimagenes[alineacionimagen]: imagen}
            self.hoja.set_header(texto, opciones)
        else:
            self.hoja.set_header(texto)

    def RepetirFila(self, primera=0, ultima=None):
        if ultima:
            self.hoja.repeat_rows(primera, ultima)
        else:
            self.hoja.repeat_rows(primera)

    def AjustarPaginas(self, ancho=1, alto=0, papel='A4'):
        """
        Ajuste el área impresa a un número específico de páginas tanto vertical como horizontalmente.
        :param ancho Número de páginas horizontalmente
        :param alto Número de páginas verticalmente (si esta en 0 se ajusta al alto necesario)
        :return:
        """
        papeles = {
            'A4': 9,
            'Oficio': 5,
        }
        self.hoja.set_paper(papeles.get(papel) or 0)
        self.hoja.fit_to_pages(ancho, alto)

    def NombreFilaColumna(self, fila=0, col=0):
        return xl_rowcol_to_cell(row=fila, col=col)

    def CombinaFilaColumna(self, desdefila=0, hastafila=0, desdecol=0, hastacol=0):
        self.hoja.merge_range(desdefila, desdecol, hastafila, hastacol, '')

    def Exporta(self, datos, cabeceras, titulo="", archivo=""):

        archivo = archivo.replace('.', '').replace('/', '')
        if not archivo.startswith("excel"):
            archivo = "excel/" + archivo

        cArchivo = self.ObtieneArchivo(archivo)
        if not cArchivo:
            return

        desde = "A"
        hasta = chr(65 + len(datos[0]))
        self.Titulo(titulo, desdecol=desde, hastacol=hasta, combina=True)

        self.EstableceEncabezado(encabezados=cabeceras)

        for row in datos:
            self.EscribeFila(row)

        self.Cerrar()

    def AbrirArchivo(self, c_archivo='', solo_datos=True, hoja=''):
        if not c_archivo.upper().endswith(('XLS', 'XLSX')):
            c_archivo += '.xlsx'

        if not os.path.isfile(c_archivo):
            Ventanas.showAlert("Sistema", f"Verifique el archivo pasado {c_archivo} no existe")
        self.wb = load_workbook(c_archivo, data_only=solo_datos)
        if self.hoja:
            self.hoja = self.wb[hoja]
        self.archivo = c_archivo
        self.max_row = self.wb.active.max_row

    def selecciona_hoja(self, hoja='Sheet1'):
        self.hoja = self.wb[hoja]

    def LeerCelda(self, fila, columna=0):
        if not self.VerificaAperturaArchivo():
            return False

        ws = self.wb.active
        if isinstance(fila, str):
            celda = ws[fila]
        else:
            celda = ws[self.NombreFilaColumna(fila, columna)]

        return celda.value if celda else ''

    def VerificaAperturaArchivo(self):
        if not self.wb:
            Ventanas.showAlert("Sistema", "No se encuentra un Archivo abierto")
            self.archivo_abierto = False
        else:
            self.archivo_abierto = True
        return self.archivo_abierto

    def NumeroColumna(self, col_str):
        return ord(col_str) - 65


@contextlib.contextmanager
def load_xl_file(xlfilepath):
    ''' Open an existing Excel file using a context manager
        `xlfilepath`: path to an existing Excel file '''

    xl = win32.DispatchEx("Excel.Application")
    wb = xl.Workbooks.Open(xlfilepath)
    try:
        yield wb
    except:
        pass
    finally:
        wb.Close(SaveChanges=True)
        xl.Quit()
        xl = None  # this actually ends the process


def xlautofit(xlfilepath, skip_first_col=False):
    ''' relies on win32com.client to autofit columns on data sheets

        remember that this is using COM so sheet numbers start at 1 (not 0),
        so to avoid requiring the caller to remember this, we increment

        returns full path (including dir) to file '''
    if os.path.splitext(xlfilepath)[1].upper() not in ('.XLS', '.XLSX'):
        raise Exception('Archivo con extension no valida: {}'.format(xlfilepath))

    autofitbegcol = 1
    if skip_first_col:
        autofitbegcol += 1

    # Autofit every sheet
    with load_xl_file(xlfilepath) as wb:
        for ws in wb.Sheets:
            autofitendcol = ws.UsedRange.Columns.Count
            ws.Range(ws.Cells(1, autofitbegcol),
                     ws.Cells(1, autofitendcol)).EntireColumn.AutoFit()
    return xlfilepath
